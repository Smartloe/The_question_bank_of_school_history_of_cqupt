import pandas as pd
import aiohttp
import asyncio
from openpyxl import load_workbook
import os

# 请求头和Cookies
headers = {
	'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0',
}
cookies = {
	'JSESSIONID': 'C883********************003C307',  # 此处填写你的JSESSIONID，在考试网页中可以通过浏览器开发者工具查看
}


# 异步获取题目
async def get_question(session):
	url = 'http://172.20.2.22:8080/index/examination/getTestQuestions'
	async with session.get(url, headers=headers, cookies=cookies, ssl=False) as response:
		return await response.json()


# 异步检查答案
async def check_answer(session, qid, answer):
	url = 'http://172.20.2.22:8080/index/examination/checkAnswer'
	data = {
		'QID': qid,
		'answer': answer,
	}
	async with session.post(url, headers=headers, cookies=cookies, data=data, ssl=False) as response:
		result = await response.json()
		return result['data'] == "1"


# 使用试错法找到正确答案并保存题目和答案
async def determine_correct_answers(session):
	questions = (await get_question(session))['data']['questionList']
	correct_answers = []

	for question in questions:
		qid = question['QID']
		qcontent = question['QContent']
		for answer in question['answers']:
			if await check_answer(session, qid, answer['AVal']):
				correct_answers.append({
					'QID': qid,
					'QContent': qcontent,
					'CorrectAnswer': answer['AText']
				})
				break

	return correct_answers


# 将结果追加保存到Excel
def append_to_excel(data, filename='output.xlsx'):
	new_data = pd.DataFrame(data)
	if os.path.exists(filename):
		# 如果文件存在，则追加数据
		book = load_workbook(filename)
		writer = pd.ExcelWriter(filename, engine='openpyxl')
		writer.book = book
		writer.sheets = {ws.title: ws for ws in book.worksheets}
		for sheetname in writer.sheets:
			new_data.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False, header=False)
		writer.save()
	else:
		# 如果文件不存在，则创建新文件
		new_data.to_excel(filename, index=False)

	print("结果已保存到", filename)


# 主执行块
async def main():
	async with aiohttp.ClientSession() as session:
		print("正在获取题目...")
		for i in range(300):  # 获取题库的次数，可以根据需要调整
			print(f"第{i + 1}次获取题目...")
			correct_answers = await determine_correct_answers(session)
			print(correct_answers)
			append_to_excel(correct_answers)
			print(f"第{i + 1}次获取题目完成！")


if __name__ == '__main__':
	asyncio.run(main())
